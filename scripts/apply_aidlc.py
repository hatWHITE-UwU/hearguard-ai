#!/usr/bin/env python3
"""
Aplica la estructura AI-DLC completa (77 columnas, 5 secciones)
a la hoja 'Matriz de registro' del Excel de HearGuard AI.
Mejoras v2: alternancia de filas, resumen final, auto-filtro, impresion.
"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from pathlib import Path

OUT = Path(__file__).resolve().parents[1] / "docs" / "matriz-registro-hearguard.xlsx"

# ── Paleta HearGuard AI ───────────────────────────────────────────────────────
BRAND   = "0D1117"   # fondo oscuro real del sistema
NAVY    = "1F4E79"   # azul marino institucional
ACCENT  = "00BCD4"   # cyan (aprox #00E5FF rebajado para Excel)
PURPLE  = "5C35CC"   # purple rebajado

# Colores por seccion (fondo cabecera)
S_BG = {
    "s1":  "1F3864",   # Identificacion
    "s2":  "1A5276",   # AI-DLC fases
    "s21": "1B6CA8",
    "s22": "1E88C7",
    "s23": "21A0DC",
    "s3":  "1B6A4A",   # Coherencia
    "s31": "1E8A5E",
    "s32": "1FA870",
    "s4":  "6B2A8B",   # Trazabilidad
    "s41": "8B3FAD",
    "s5":  "8B2500",   # Estabilidad
    "s51": "A93000",
    "s52": "C43A00",
    "sf":  "2C2C2C",   # Final
}

# Colores por intent (fila principal / fila alternada)
INT_MAIN = {
    "INT-001": "D6EAD6", "INT-002": "D6E8F5", "INT-003": "FFF0C0",
    "INT-004": "FAD7C5", "INT-005": "F8D0D0", "INT-006": "E0D6F5",
    "INT-007": "D0DCF5", "INT-008": "F9D8B8", "INT-009": "CCE0FF",
    "INT-010": "CCF0DA", "INT-011": "FFF0B8", "INT-012": "EDD6FF",
    "INT-013": "D6F0FA", "INT-014": "FADDDD", "INT-015": "D6F5D6",
}
INT_ALT = {
    "INT-001": "EAF4EA", "INT-002": "EAF3FA", "INT-003": "FFFAE0",
    "INT-004": "FDE8DC", "INT-005": "FCE8E8", "INT-006": "EDE8FA",
    "INT-007": "E4ECFA", "INT-008": "FCEAD8", "INT-009": "E0EEFF",
    "INT-010": "E0F8E8", "INT-011": "FFFFF0", "INT-012": "F5E8FF",
    "INT-013": "E8F7FD", "INT-014": "FDE8E8", "INT-015": "E8FAE8",
}
INT_LABEL = {
    "INT-001": "276221", "INT-002": "1F4E79", "INT-003": "7D5A00",
    "INT-004": "843C1C", "INT-005": "9C0006", "INT-006": "4A1F8C",
    "INT-007": "1A3A6E", "INT-008": "7D4510", "INT-009": "0A2F6E",
    "INT-010": "1A5C2E", "INT-011": "6E5A00", "INT-012": "4A0F7A",
    "INT-013": "0A4A6E", "INT-014": "7A0A0A", "INT-015": "0A5C0A",
}

# Semaforo de resultado
RES_FILL = {
    "Aprobado":   "C6EFCE", "Sí":        "C6EFCE",
    "En progreso":"FFEB9C", "Parcial":   "FFEB9C",
    "Pendiente":  "FFC7CE", "No":        "FFC7CE",
    "N/A":        "EDEDED",
}
RES_FONT = {
    "Aprobado":   "276221", "Sí":        "276221",
    "En progreso":"7D4B12", "Parcial":   "7D4B12",
    "Pendiente":  "9C0006", "No":        "9C0006",
    "N/A":        "555555",
}

RESP_LUIS  = "Terreros H. L.F."
RESP_HARDY = "Rondinel A. H.E."

# ── Utilidades de estilo ─────────────────────────────────────────────────────
def _fill(h): return PatternFill("solid", fgColor=h)

def _border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, col_s, col_e, txt, bg, fc="FFFFFF", sz=9, bold=True):
    if col_s != col_e:
        ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_e)
    c = ws.cell(row, col_s, value=txt)
    c.fill = _fill(bg); c.font = Font(bold=bold, size=sz, color=fc)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    # fill merged right cells border
    for col in range(col_s+1, col_e+1):
        ws.cell(row, col).border = _border()

def _status(cell, val, sz=8):
    cell.value = val
    cell.fill  = _fill(RES_FILL.get(val, "FFFFFF"))
    cell.font  = Font(bold=True, size=sz, color=RES_FONT.get(val, "000000"))
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

# ── Datos de bolts ────────────────────────────────────────────────────────────
RAW = [
# (iid, title, num, desc, modelo, fecha, dominio, hu, resp, status, ev, notes, mejora)
("INT-001","Framework de testing (TDD)","001","Configurar Jest+Supertest+lcov umbral CI 100%","Claude Sonnet 4.6","03/2026","Backend","RF-01–06",RESP_LUIS,"Si","backend/package.json · jest.config.js","Base de calidad TDD",""),
("INT-001","Framework de testing (TDD)","002","Suite integración: auth,noise,evaluation,device,middleware (207 tests)","Claude Sonnet 4.6","03/2026","Backend","RF-01–06",RESP_LUIS,"Si","backend/tests/ — 207 tests","",""),
("INT-001","Framework de testing (TDD)","003","Tests de seguridad: JWT,IDOR,NoSQL injection,rutas (22 casos)","Claude Sonnet 4.6","03/2026","Seguridad","RF-01,RF-06",RESP_LUIS,"Si","security.test.js — 22 casos","",""),
("INT-001","Framework de testing (TDD)","004","pytest + coverage XML en ai-service (30 tests)","Claude Sonnet 4.6","03/2026","IA","RF-04",RESP_LUIS,"Si","ai-service/tests/ — 30 tests","",""),
("INT-001","Framework de testing (TDD)","005","Vitest Angular: core + features services (107 tests)","Claude Sonnet 4.6","03/2026","Frontend","RF-01,RF-03",RESP_LUIS,"Si","frontend/**/*.spec.ts — 107 tests","",""),
("INT-001","Framework de testing (TDD)","006","flutter_test movil — 4 archivos (42 tests)","Claude Sonnet 4.6","03/2026","Movil","—",RESP_LUIS,"Si","flutter_app/test/ — 42 tests","",""),
("INT-001","Framework de testing (TDD)","007","Playwright E2E smoke+auth+hearing-test (36 tests)","Claude Sonnet 4.6","03/2026","E2E","RF-05",RESP_LUIS,"Si","e2e/tests/ — 36 tests","Vercel preview",""),
("INT-001","Framework de testing (TDD)","008","Plan de pruebas IEEE 829 / ISO 29119 (1169 lineas)","Claude Sonnet 4.6","03/2026","Docs","—",RESP_LUIS,"Si","docs/plan-de-pruebas.md","1169 lineas",""),
("INT-001","Framework de testing (TDD)","009","coverage-extra + noise.service + database tests — 100% ramas","Claude Sonnet 4.6","21/05/2026","Backend","—",RESP_LUIS,"Si","Commit 8ebc768","100% branches",""),
("INT-001","Framework de testing (TDD)","010","evaluation-ai.test.js — flujo IA exitoso con mock","Claude Sonnet 4.6","22/05/2026","Backend","RF-04",RESP_LUIS,"Si","Commit f8daa25","",""),
("INT-001","Framework de testing (TDD)","011","Jest --runInBand — estabilidad MongoDB en cobertura","Equipo","22/05/2026","Backend","—",RESP_LUIS,"Si","README seccion Tests","Sin --runInBand: race condition",""),
# INT-002
("INT-002","BDD — Gherkin y trazabilidad","001","6 archivos .feature (auth,ruido,auditiva,IA,IoT,resultados) — 85 escenarios","Claude Sonnet 4.6","03/2026","BDD","RF-01–06",RESP_LUIS,"Si","docs/features/ — 85 escenarios","",""),
("INT-002","BDD — Gherkin y trazabilidad","002","Matriz trazabilidad RF/RNF <-> BDD <-> test (60 RF, 10 RNF)","Claude Sonnet 4.6","03/2026","Docs","—",RESP_LUIS,"Si","docs/matriz-trazabilidad.md","",""),
("INT-002","BDD — Gherkin y trazabilidad","003","Metodologia TDD+BDD documentada con referencias APA","Claude Sonnet 4.6","05/2026","Docs","—",RESP_LUIS,"Si","docs/metodologia.md","",""),
("INT-002","BDD — Gherkin y trazabilidad","004","Ejecutar .feature con Cucumber en CI","Claude Sonnet 4.6","06/2026","BDD","—",RESP_LUIS,"Si","job bdd en ci.yml — 85 escenarios Cucumber.js","",""),
# INT-003
("INT-003","CI/CD — Validaciones automaticas","001","6 jobs: backend,ai-service,frontend,e2e,flutter,sonarcloud","Claude Sonnet 4.6","03–05/2026","CI/CD","RNF-09",RESP_LUIS,"Si","ci.yml — 6 jobs","",""),
("INT-003","CI/CD — Validaciones automaticas","002","Artefactos cobertura + fix-sonar-coverage-paths.js","Claude Sonnet 4.6","21/05/2026","CI/CD","RNF-03",RESP_LUIS,"Si","scripts/fix-sonar-coverage-paths.js","",""),
("INT-003","CI/CD — Validaciones automaticas","003","deploy.yml — Render (backend+IA) + Vercel (frontend)","Claude Sonnet 4.6","03/2026","Deploy","—",RESP_LUIS,"Si","deploy.yml","",""),
("INT-003","CI/CD — Validaciones automaticas","004","Badge CI verde en README","GitHub Actions","05/2026","CI/CD","—",RESP_LUIS,"Si","README badge CI","",""),
("INT-003","CI/CD — Validaciones automaticas","005","Optimizar duracion pipeline con metricas SLA","—","—","CI/CD","—",RESP_LUIS,"Parcial","Jobs paralelos OK","Sin benchmark formal","Agregar job de metricas SLA"),
# INT-004
("INT-004","Calidad — SonarCloud","001","sonar-project.properties + exclusiones CPD y cobertura","Claude Sonnet 4.6","05/2026","Calidad","RNF-09",RESP_LUIS,"Si","sonar-project.properties","",""),
("INT-004","Calidad — SonarCloud","002","Ratings Security/Reliability/Maintainability = A","Claude Sonnet 4.6","21/05/2026","Calidad","—",RESP_LUIS,"Si","Commit 28925d5","0 issues abiertos",""),
("INT-004","Calidad — SonarCloud","003","Cerrar S2699 y 37+ code smells de mantenibilidad","Claude Sonnet 4.6","20–21/05/2026","Calidad","—",RESP_LUIS,"Si","Commits ddde986,319e9bc","",""),
("INT-004","Calidad — SonarCloud","004","Duplicacion 0% + cobertura 100% en SonarCloud","Claude Sonnet 4.6","22/05/2026","Calidad","RNF-03",RESP_LUIS,"Si","Commit 8ebc768","",""),
# INT-005
("INT-005","Seguridad","001","S5147 NoSQL injection: $eq en Device,Noise,buildDateFilter","Claude Sonnet 4.6","19–20/05/2026","Seguridad","RNF-06",RESP_LUIS,"Si","Commits 25c1603,1998062","",""),
("INT-005","Seguridad","002","S2068 + PRNG seguro en helpers E2E (crypto.randomBytes)","Claude Sonnet 4.6","20/05/2026","Seguridad","—",RESP_LUIS,"Si","Commits fb4119c,8c5d890","Regresion Rating C->A",""),
("INT-005","Seguridad","003","npm audit periodico — sin vulnerabilidades criticas","Equipo","—","Seguridad","—",RESP_LUIS,"Parcial","Manual — npm audit OK","Sin job automatico bundler-audit","Agregar job npm audit en CI"),
# INT-006
("INT-006","Modelo IA — CRISP-DM","001","CRISP-DM Fase 1-2: comprension negocio y variables auditivas","Claude Sonnet 4.6","03/2026","IA","RF-04",RESP_LUIS,"Si","README · docs/metodologia.md","",""),
("INT-006","Modelo IA — CRISP-DM","002","Fase 3: features.py + constants.py (8 features tonales)","Claude Sonnet 4.6","03/2026","IA","RF-04",RESP_LUIS,"Si","ai-service/model/features.py","8 features 250-8000 Hz",""),
("INT-006","Modelo IA — CRISP-DM","003","Fase 4: trainer.py RandomForest SEED=42 (5000 muestras)","Claude Sonnet 4.6","03/2026","IA","RF-04",RESP_LUIS,"Si","ai-service/model/trainer.py","n_estimators=120 depth=12",""),
("INT-006","Modelo IA — CRISP-DM","004","Fase 5: R2>=0.80 holdout — perfiles bajo/alto validados","Claude Sonnet 4.6","03/2026","IA","RF-04",RESP_LUIS,"Si","ai-service/tests/test_predictor.py","R2 holdout ~0.85",""),
("INT-006","Modelo IA — CRISP-DM","005","Fase 6: Flask /predict-risk + /recommend + ai.service.js","Claude Sonnet 4.6","03/2026","IA","RF-04",RESP_LUIS,"Si","app.py · backend/src/services/ai.service.js","Degradacion elegante",""),
("INT-006","Modelo IA — CRISP-DM","006","Reentrenamiento deterministico en CI (job ai-service)","Claude Sonnet 4.6","05/2026","IA","—",RESP_LUIS,"Si","ci.yml job ai-service","SEED=42 reproducible",""),
# INT-007
("INT-007","Rendimiento (RNF)","001","k6: smoke,load,spike — umbrales p95<2s y error<5%","Claude Sonnet 4.6","03/2026","Rendimiento","RNF-01,02",RESP_LUIS,"Si","tests/k6/load-test.js","3 escenarios",""),
("INT-007","Rendimiento (RNF)","002","Ejecutar k6 contra Render produccion y adjuntar reporte","Claude Sonnet 4.6","06/2026","Rendimiento","RNF-01",RESP_LUIS,"Si","k6-smoke job ci.yml — reports/k6/index.html (handleSummary)","",""),
("INT-007","Rendimiento (RNF)","003","Lighthouse en frontend Vercel (Performance,A11y,SEO)","Claude Sonnet 4.6","06/2026","Frontend","—",RESP_LUIS,"Si","lighthouse job ci.yml (treosh/lighthouse-ci-action) — accessibility >=90%","",""),
# INT-008
("INT-008","Documentacion y entregables","001","articulo.md + README.md + api-spec.yml OpenAPI 3.1","Claude Sonnet 4.6","05/2026","Docs","—",RESP_LUIS,"Si","docs/articulo.md · README.md · docs/api-spec.yml","",""),
("INT-008","Documentacion y entregables","002","Matriz de registro Excel AI-DLC v3 (77 columnas)","Claude Sonnet 4.6","06/2026","Docs","—",RESP_LUIS,"Si","docs/matriz-registro-hearguard.xlsx","v3.0",""),
("INT-008","Documentacion y entregables","003","complejidad-ciclomatica.md (McCabe — 17 funciones)","Claude Sonnet 4.6","03/2026","Docs","—",RESP_LUIS,"Si","docs/complejidad-ciclomatica.md","59 caminos independientes",""),
("INT-008","Documentacion y entregables","004","Docker Compose reproducible — stack completo local","Claude Sonnet 4.6","03/2026","DevOps","—",RESP_LUIS,"Si","docker-compose.yml","",""),
# INT-009
("INT-009","API REST — Autenticacion","001","POST /api/auth/register — JWT access+refresh+bcrypt salt=12","Claude Sonnet 4.6","08/2025","Backend","RF-01",RESP_LUIS,"Si","auth.controller.js:register · auth.test.js","",""),
("INT-009","API REST — Autenticacion","002","POST /api/auth/login — anti-timing oracle BCRYPT_DUMMY","Claude Sonnet 4.6","08/2025","Backend","RF-01",RESP_LUIS,"Si","auth.controller.js:login · security.test.js","S2-003 passed",""),
("INT-009","API REST — Autenticacion","003","POST /api/auth/refresh — rotacion SHA-256 timingSafeEqual","Claude Sonnet 4.6","08/2025","Backend","RF-01",RESP_LUIS,"Si","auth.controller.js:refresh","rotacion en cada uso",""),
("INT-009","API REST — Autenticacion","004","POST /api/auth/logout — revocacion hash refreshToken","Claude Sonnet 4.6","08/2025","Backend","RF-01",RESP_LUIS,"Si","auth.controller.js:logout","hash=null en BD",""),
("INT-009","API REST — Autenticacion","005","GET /api/auth/me — perfil del usuario autenticado","Claude Sonnet 4.6","08/2025","Backend","RF-01",RESP_LUIS,"Si","auth.controller.js:me","",""),
("INT-009","API REST — Autenticacion","006","PATCH /api/auth/me — actualizacion nombre,edad,settings","Claude Sonnet 4.6","09/2025","Backend","RF-01",RESP_LUIS,"Si","auth.controller.js:patchMe","isPlainObject guard",""),
# INT-010
("INT-010","API REST — Monitoreo de Ruido","001","POST /api/noise — registro lectura + deviceId opcional","Claude Sonnet 4.6","09/2025","Backend","RF-02",RESP_LUIS,"Si","noise.controller.js:create · noise.test.js","",""),
("INT-010","API REST — Monitoreo de Ruido","002","POST /api/noise/iot — X-Device-Key CSPRNG autenticacion","Claude Sonnet 4.6","09/2025","Backend","RF-02,RF-06",RESP_LUIS,"Si","noise.controller.js:createIot","",""),
("INT-010","API REST — Monitoreo de Ruido","003","GET /api/noise — historial filtros fecha,source,skip,limit","Claude Sonnet 4.6","09/2025","Backend","RF-02",RESP_LUIS,"Si","noise.controller.js:list · buildDateFilter","NoSQL-safe",""),
("INT-010","API REST — Monitoreo de Ruido","004","GET /api/noise/stats/today — exposicion diaria dB","Claude Sonnet 4.6","10/2025","Backend","RF-02",RESP_LUIS,"Si","noise.service.js:statsForToday","",""),
("INT-010","API REST — Monitoreo de Ruido","005","GET /api/noise/stats/week — exposicion 7 dias","Claude Sonnet 4.6","10/2025","Backend","RF-02",RESP_LUIS,"Si","noise.service.js:statsForWeek","",""),
# INT-011
("INT-011","API REST — Evaluacion Auditiva","001","POST /api/evaluations — crear eval + IA + degradacion elegante","Claude Sonnet 4.6","10/2025","Backend","RF-03,RF-04",RESP_LUIS,"Si","evaluation.controller.js:create · evaluation-ai.test.js","riskResult=null si IA falla",""),
("INT-011","API REST — Evaluacion Auditiva","002","GET /api/evaluations — historial paginado limit,skip","Claude Sonnet 4.6","10/2025","Backend","RF-03",RESP_LUIS,"Si","evaluation.controller.js:list","",""),
("INT-011","API REST — Evaluacion Auditiva","003","GET /api/evaluations/:id — detalle con riskResult embebido","Claude Sonnet 4.6","10/2025","Backend","RF-03,RF-05",RESP_LUIS,"Si","evaluation.controller.js:getById","",""),
("INT-011","API REST — Evaluacion Auditiva","004","PATCH /api/evaluations/:id — habitData,status,frequencyScores","Claude Sonnet 4.6","11/2025","Backend","RF-03",RESP_LUIS,"Si","evaluation.controller.js:patch","whitelist allowed keys",""),
# INT-012
("INT-012","API REST — Dispositivos IoT","001","POST /api/devices — registro con apiKey CSPRNG 32 bytes","Claude Sonnet 4.6","11/2025","Backend","RF-06",RESP_LUIS,"Si","device.controller.js · device.test.js","crypto.randomBytes",""),
("INT-012","API REST — Dispositivos IoT","002","GET /api/devices — listar dispositivos del usuario auth","Claude Sonnet 4.6","11/2025","Backend","RF-06",RESP_LUIS,"Si","device.controller.js:list","",""),
# INT-013
("INT-013","Frontend Angular 17+","001","SPA standalone + Signals — 13 componentes funcionales","Claude Sonnet 4.6","10–11/2025","Frontend","RF-01–05",RESP_LUIS,"Si","frontend/src/app/features/ — 13 componentes","",""),
("INT-013","Frontend Angular 17+","002","Web Audio API — captura RMS->dBSPL + histograma 30 muestras","Claude Sonnet 4.6","11/2025","Frontend","RF-02",RESP_LUIS,"Si","noise-monitor.service.ts:sample","formula 20*log10(rms+1e-7)+94",""),
("INT-013","Frontend Angular 17+","003","authInterceptor — JWT automatico + refresco silencioso 401","Claude Sonnet 4.6","11/2025","Frontend","RF-01",RESP_LUIS,"Si","auth.interceptor.ts · auth.interceptor.spec.ts","107 tests incluyen interceptor",""),
("INT-013","Frontend Angular 17+","004","Design system oscuro: #0D1117,#00E5FF,#7C4DFF + Poppins","Claude Sonnet 4.6","10/2025","Frontend","—",RESP_LUIS,"Si","styles.scss · environments/","",""),
("INT-013","Frontend Angular 17+","005","auth.guard + rutas protegidas + redirect a /login","Claude Sonnet 4.6","11/2025","Frontend","RF-01",RESP_LUIS,"Si","auth.guard.ts · auth.guard.spec.ts","",""),
# INT-014
("INT-014","App Movil Flutter 3","001","10 pantallas nativas (splash,auth,dashboard,monitor,hearing,results,history,profile)","Claude Sonnet 4.6","12/2025","Movil","RF-01–05",RESP_LUIS,"Si","flutter_app/lib/features/ — 10 pantallas","",""),
("INT-014","App Movil Flutter 3","002","noise_meter — captura microfono real + permission_handler runtime","Claude Sonnet 4.6","12/2025","Movil","RF-02",RESP_LUIS,"Si","monitor_screen.dart:_startListening","permission en runtime",""),
("INT-014","App Movil Flutter 3","003","SharedPreferences + AuthService ChangeNotifier + Provider","Claude Sonnet 4.6","12/2025","Movil","RF-01",RESP_LUIS,"Si","auth_service.dart — 42 tests","",""),
("INT-014","App Movil Flutter 3","004","ApiClient Dio — HTTP client tokens + interceptors movil","Claude Sonnet 4.6","12/2025","Movil","RF-01",RESP_LUIS,"Si","api_client.dart","",""),
# INT-015
("INT-015","IoT Firmware","001","Firmware ESP32 WiFi+HTTP — KY-037 -> POST /api/noise/iot","Claude Sonnet 4.6","01/2026","IoT","RF-06",RESP_LUIS,"Si","arduino/hearguard_esp32/hearguard_esp32.ino","",""),
("INT-015","IoT Firmware","002","Arduino Uno serial + serial_bridge.js -> backend Node.js","Claude Sonnet 4.6","01/2026","IoT","RF-06",RESP_LUIS,"Si","arduino/serial_bridge.js","DEVICE_KEY env var",""),
("INT-015","IoT Firmware","003","Simulacion Wokwi ESP32 (diagram.json, sketch.ino) sin hardware","Claude Sonnet 4.6","01/2026","IoT","RF-06",RESP_LUIS,"Si","arduino/wokwi/","Sin hardware fisico",""),
]

def _build(raw):
    iid,ititle,bnum,desc,modelo,fecha,dominio,hu,resp,status,ev,notes,mejora = raw
    bid   = f"{iid}-BOLT-{bnum}"
    is_ok = status in ("Si","Si")
    is_par= status == "Parcial"

    def r(v_ok, v_par="Parcial", v_no="No"):
        if is_ok:  return v_ok
        if is_par: return v_par
        return v_no

    deploy_r = r("Si","Parcial","No")
    deploy_e = r("Render / Vercel / CI","CI parcial","N/A — no implementado")

    return {
        "iid":iid,"ititle":ititle,"bid":bid,"desc":desc,"modelo":modelo,
        "fecha":fecha,"dominio":dominio,"hu":hu,"resp":resp,
        "status":status,"ev":ev,"notes":notes,"mejora":mejora,
        # S2.1
        "s21ir":"Si", "s21ie":"Matriz de registro",
        "s21rr":"Si", "s21re":"docs/plan-de-pruebas.md",
        "s21br":"Si", "s21be":bid, "s21x":"Aprobado",
        # S2.2
        "s22dr":r("Si","Si","No"), "s22de":r("README · docs/api-spec.yml","Parcialmente","No implementado"),
        "s22cr":r("Si","Parcial","No"), "s22ce":ev,
        "s22pr":r("Si","Parcial","No"), "s22pe":r("507 tests pasando","Tests parciales","Sin tests"),
        "s22x":r("Aprobado","En progreso","Pendiente"),
        # S2.3
        "s23dr":deploy_r, "s23de":deploy_e,
        "s23x":r("Aprobado","En progreso","Pendiente"),
        "s2x":r("Aprobado","En progreso","Pendiente"),
        # S3.1
        "s31mr":r("Si","Si","N/A"), "s31me":r(ev,"—","—"),
        "s31cr":r("Si","Si","N/A"), "s31ce":r(f"MVC / {dominio}","—","—"),
        "s31rr":r("Si","Si","N/A"), "s31re":r("Sin duplicacion","—","—"),
        "s31sr":r("Si","Si","N/A"), "s31se":r("Un modulo por archivo","—","—"),
        "s31x":r("Aprobado","Aprobado","N/A"),
        # S3.2
        "s32nr":r("Si","Si","N/A"), "s32ne":r("camelCase JS / snake_case PY","—","—"),
        "s32or":r("Si","Si","N/A"), "s32oe":r("Estructura MVC / carpetas","—","—"),
        "s32dr":r("Si","Si","N/A"), "s32de":r("Librerias estandar del proyecto","—","—"),
        "s32x":r("Aprobado","Aprobado","N/A"),
        "s3x":r("Aprobado","Aprobado","N/A"),
        # S4.1
        "s41br":"Si", "s41be":bid,
        "s41rr":"Si", "s41re":hu or "—",
        "s41cr":r("Si","Parcial","No"), "s41ce":ev,
        "s41vr":r("Si","Parcial","No"), "s41ve":r("Tests automatizados","Validacion manual","Sin validacion"),
        "s41pr":r("Si","No","No"),     "s41pe":r("Render/Vercel/CI","N/A","N/A"),
        "s4x":r("Aprobado","En progreso","Pendiente"),
        # S5.1
        "s51fr":r("Si","Parcial","No"), "s51fe":r("Tests pasando","Func. parcial","No implementado"),
        "s51ar":r("Si","Parcial","No"), "s51ae":r("Cobertura 100%","Alcance limitado","—"),
        "s51pr":r("Si","Si","N/A"),    "s51pe":r("Sin regresiones","Sin regresiones","—"),
        "s51ir":r("Si","Si","N/A"),    "s51ie":r("Sin correccion inmediata","—","—"),
        "s51x":r("Aprobado","En progreso","Pendiente"),
        # S5.2
        "s52sr":r("Si","Si","N/A"), "s52se":r("ESLint 0 err · SonarCloud A","ESLint OK","—"),
        "s52cr":r("Si","Si","N/A"), "s52ce":r("npm audit · security.test.js","npm audit OK","—"),
        "s52pr":r("Si","Parcial","No"), "s52pe":r("507/507 tests","Tests parciales","Sin tests"),
        "s52rr":r("Si","Parcial","No"), "s52re":r("CI pipeline verde","Revision parcial","Sin revision"),
        "s52x":r("Aprobado","En progreso","Pendiente"),
        "s5x":r("Aprobado","En progreso","Pendiente"),
        # Final
        "gx":r("Aprobado","En progreso","Pendiente"),
        "estado":r("Completado","En progreso","Pendiente"),
        "obs":notes, "mejora":mejora or "—",
    }

# ── Calcular posicion alternada dentro de cada intent ─────────────────────────
BOLTS = [_build(r) for r in RAW]
_cnt  = {}
for b in BOLTS:
    _cnt.setdefault(b["iid"], 0)
    b["_alt"] = _cnt[b["iid"]] % 2
    _cnt[b["iid"]] += 1

# ── Columnas de resultado (para formatos condicionales y ancho) ───────────────
RESULT_COLS = {15,22,25,26,35,42,43,54,63,72,73,74,75}

# ── Constructor de la hoja ────────────────────────────────────────────────────
def write_matriz(ws):
    ws.sheet_view.showGridLines = False
    MAX_COL = 77

    def cl(n): return get_column_letter(n)

    # ── Fila 1: Titulo ──
    ws.merge_cells(f"A1:{cl(MAX_COL)}1")
    c = ws["A1"]
    c.value = "HearGuard AI  |  Matriz de Registro de Bolts — Marco AI-DLC  |  v3.0"
    c.font  = Font(bold=True, size=13, color="FFFFFF")
    c.fill  = _fill("0D1117")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Fila 2: Autores y metodologia ──
    ws.merge_cells(f"A2:{cl(MAX_COL)}2")
    c = ws["A2"]
    c.value = ("Terreros Hinojosa Luis Francisco (autor principal)  ·  Rondinel Aquino Hardy Eduardo (coautor)  ·  "
               "Universidad Continental  ·  Metodologia: TDD+BDD + CRISP-DM")
    c.font  = Font(italic=True, size=8, color="CCDDFF")
    c.fill  = _fill("1A2540")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 15

    # ── Fila 3: Secciones principales ──
    _hdr(ws,3, 1,  8, "1. Identificacion",                              S_BG["s1"],  sz=9)
    _hdr(ws,3, 9, 26, "2. Registro de la aplicacion del marco AI-DLC",  S_BG["s2"],  sz=9)
    _hdr(ws,3,27, 43, "3. Registro de coherencia tecnica",               S_BG["s3"],  sz=9)
    _hdr(ws,3,44, 54, "4. Registro de trazabilidad",                     S_BG["s4"],  sz=9)
    _hdr(ws,3,55, 73, "5. Registro de estabilidad",                      S_BG["s5"],  sz=9)
    _hdr(ws,3,74, 77, "Resultado final",                                  S_BG["sf"], sz=9)
    ws.row_dimensions[3].height = 20

    # ── Fila 4: Subsecciones ──
    s1_sub = ["Intent","ID del Bolt","Descripcion del Bolt","Modelo IA usado","Fecha","Dominio funcional","HU / Ticket","Responsable"]
    for i,lbl in enumerate(s1_sub,1):
        _hdr(ws,4,i,i,lbl,S_BG["s1"],sz=8)
    _hdr(ws,4, 9,15,"2.1 Fase de Inicio",             S_BG["s21"],sz=8)
    _hdr(ws,4,16,22,"2.2 Fase de construccion",        S_BG["s22"],sz=8)
    _hdr(ws,4,23,25,"2.3 Fase de operaciones",          S_BG["s23"],sz=8)
    _hdr(ws,4,26,26,"Resultado\nSec.2",                 S_BG["s2"],sz=8)
    _hdr(ws,4,27,35,"3.1 Coherencia estructural del codigo",       S_BG["s31"],sz=8)
    _hdr(ws,4,36,42,"3.2 Consistencia con convenciones",           S_BG["s32"],sz=8)
    _hdr(ws,4,43,43,"Resultado\nSec.3",                             S_BG["s3"],sz=8)
    _hdr(ws,4,44,53,"4.1 Auditabilidad y procedencia",              S_BG["s41"],sz=8)
    _hdr(ws,4,54,54,"Resultado\nSec.4",                             S_BG["s4"],sz=8)
    _hdr(ws,4,55,63,"5.1 Estabilidad funcional del bolt",           S_BG["s51"],sz=8)
    _hdr(ws,4,64,72,"5.2 Estabilidad tecnica de integracion",       S_BG["s52"],sz=8)
    _hdr(ws,4,73,73,"Resultado\nSec.5",                             S_BG["s5"],sz=8)
    _hdr(ws,4,74,74,"Resultado\nGlobal",  S_BG["sf"],sz=8)
    _hdr(ws,4,75,75,"Estado\ndel Bolt",   S_BG["sf"],sz=8)
    _hdr(ws,4,76,76,"Observaciones\ncomplementarias",S_BG["sf"],sz=8)
    _hdr(ws,4,77,77,"Punto\nde mejora",   S_BG["sf"],sz=8)
    ws.row_dimensions[4].height = 28

    # ── Fila 5: Preguntas (merged en pares Reg/Ev) ──
    def q2(cs, txt, bg):
        _hdr(ws,5,cs,cs+1,txt,bg,sz=7)
    def q1(c, txt, bg):
        _hdr(ws,5,c,c,txt,bg,sz=7)

    for i in range(1,9): q1(i,"",S_BG["s1"])
    q2( 9,"Intent identificado?",                    S_BG["s21"])
    q2(11,"Requerimiento asociado documentado?",      S_BG["s21"])
    q2(13,"Bolt derivado del requerimiento?",         S_BG["s21"])
    q1(15,"Resultado\n2.1",                          S_BG["s21"])
    q2(16,"Diseno documentado?",                     S_BG["s22"])
    q2(18,"Implementacion registrada en commit?",    S_BG["s22"])
    q2(20,"Pruebas automatizadas ejecutadas?",       S_BG["s22"])
    q1(22,"Resultado\n2.2",                          S_BG["s22"])
    q2(23,"Evidencia de despliegue en produccion?",  S_BG["s23"])
    q1(25,"Resultado\n2.3",                          S_BG["s23"])
    q1(26,"Resultado\nSec.2",                        S_BG["s2"])
    q2(27,"Codigo en modulo correspondiente?",       S_BG["s31"])
    q2(29,"Logica en capa adecuada?",                S_BG["s31"])
    q2(31,"Se evita creacion de estructuras redundantes?", S_BG["s31"])
    q2(33,"Separacion clara de responsabilidades?",  S_BG["s31"])
    q1(35,"Resultado\n3.1",                          S_BG["s31"])
    q2(36,"Uso consistente de convenciones de nombres?", S_BG["s32"])
    q2(38,"Consistencia con organizacion interna archivos?", S_BG["s32"])
    q2(40,"Se evita dependencias innecesarias?",     S_BG["s32"])
    q1(42,"Resultado\n3.2",                          S_BG["s32"])
    q1(43,"Resultado\nSec.3",                        S_BG["s3"])
    q2(44,"Bolt documentado con ID verificable?",    S_BG["s41"])
    q2(46,"Requerimiento con ID verificable?",       S_BG["s41"])
    q2(48,"Pull request o commit asociado?",         S_BG["s41"])
    q2(50,"Evidencia de validacion recuperable?",    S_BG["s41"])
    q2(52,"Evidencia despliegue o N/A justificado?", S_BG["s41"])
    q1(54,"Resultado\nSec.4",                        S_BG["s4"])
    q2(55,"El bolt mantiene el funcionamiento esperado?",     S_BG["s51"])
    q2(57,"El bolt funciona dentro de su alcance?",           S_BG["s51"])
    q2(59,"Preserva el funcionamiento de funcionalidades existentes?", S_BG["s51"])
    q2(61,"Se integra sin correccion correctiva inmediata?",  S_BG["s51"])
    q1(63,"Resultado\n5.1",                          S_BG["s51"])
    q2(64,"Supera validacion de analisis estatico?", S_BG["s52"])
    q2(66,"Supera validacion de seguridad?",         S_BG["s52"])
    q2(68,"Supera las pruebas automatizadas?",       S_BG["s52"])
    q2(70,"Revision de codigo aprobada en CI?",      S_BG["s52"])
    q1(72,"Resultado\n5.2",                          S_BG["s52"])
    q1(73,"Resultado\nSec.5",                        S_BG["s5"])
    for i in range(74,78): q1(i,"",S_BG["sf"])
    ws.row_dimensions[5].height = 34

    # ── Fila 6: Registro / Evidencia ──
    def re(col_r, bg="D6E4F0"):
        for col,lbl in [(col_r,"Registro"),(col_r+1,"Evidencia")]:
            c=ws.cell(6,col,value=lbl)
            c.font=Font(bold=True,size=7,color="1F3864")
            c.fill=_fill(bg); c.border=_border(); c.alignment=Alignment(horizontal="center",vertical="center")
    def rl(col, lbl="Resultado", bg="D6E4F0"):
        c=ws.cell(6,col,value=lbl)
        c.font=Font(bold=True,size=7,color="1F3864")
        c.fill=_fill(bg); c.border=_border(); c.alignment=Alignment(horizontal="center",vertical="center")

    s1_lbl=["Intent","ID del Bolt","Descripcion","Modelo IA","Fecha","Dominio","HU/Ticket","Responsable"]
    for i,lbl in enumerate(s1_lbl,1):
        c=ws.cell(6,i,value=lbl); c.font=Font(bold=True,size=7,color="FFFFFF")
        c.fill=_fill(S_BG["s1"]); c.border=_border(); c.alignment=Alignment(horizontal="center",vertical="center")

    re(9);re(11);re(13); rl(15)
    re(16);re(18);re(20); rl(22)
    re(23); rl(25); rl(26)
    re(27,"D5E8D4");re(29,"D5E8D4");re(31,"D5E8D4");re(33,"D5E8D4"); rl(35,"D5E8D4")
    re(36,"D5E8D4");re(38,"D5E8D4");re(40,"D5E8D4"); rl(42,"D5E8D4"); rl(43,"D5E8D4")
    re(44,"EBD5F5");re(46,"EBD5F5");re(48,"EBD5F5");re(50,"EBD5F5");re(52,"EBD5F5"); rl(54,"EBD5F5")
    re(55,"F8DDD5");re(57,"F8DDD5");re(59,"F8DDD5");re(61,"F8DDD5"); rl(63,"F8DDD5")
    re(64,"F8DDD5");re(66,"F8DDD5");re(68,"F8DDD5");re(70,"F8DDD5"); rl(72,"F8DDD5"); rl(73,"F8DDD5")
    for i in range(74,78): rl(i,"—","E8E8E8")
    ws.row_dimensions[6].height = 14

    # ── Filas de datos ────────────────────────────────────────────────────────
    DATA_START = 7
    wrap   = Alignment(wrap_text=True, vertical="center")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def wre(col_r, key_r, key_e, bg_r="F8F9FA"):
        vr = b.get(key_r,"")
        ve = b.get(key_e,"")
        cr = ws.cell(row, col_r)
        ce = ws.cell(row, col_r+1)
        sfill = RES_FILL.get(vr, "F0F4FF")
        sfont = RES_FONT.get(vr, "1F1F1F")
        cr.value=vr; cr.fill=_fill(sfill); cr.font=Font(bold=True,size=7,color=sfont)
        cr.alignment=center; cr.border=_border()
        ce.value=ve; ce.fill=_fill(bg_r); ce.font=Font(size=7)
        ce.alignment=wrap; ce.border=_border()

    def wres(col, key, sz=8):
        _status(ws.cell(row, col), b.get(key,""), sz)

    for i, b in enumerate(BOLTS):
        row = DATA_START + i
        # Alternancia: color principal / color claro
        row_color = INT_ALT[b["iid"]] if b["_alt"] else INT_MAIN[b["iid"]]
        lbl_color = INT_LABEL[b["iid"]]

        # S1
        c1=ws.cell(row,1,value=f"{b['iid']}\n{b['ititle']}")
        c1.fill=_fill(INT_MAIN[b["iid"]]); c1.font=Font(bold=True,size=8,color=lbl_color)
        c1.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c1.border=_border()

        s1_data=[b["bid"],b["desc"],b["modelo"],b["fecha"],b["dominio"],b["hu"],b["resp"]]
        for ci,val in enumerate(s1_data,2):
            cell=ws.cell(row,ci,value=val)
            cell.fill=_fill(row_color); cell.font=Font(size=8 if ci==3 else 8,color=lbl_color)
            cell.alignment=wrap if ci==3 else center; cell.border=_border()

        # S2.1
        wre( 9,"s21ir","s21ie"); wre(11,"s21rr","s21re"); wre(13,"s21br","s21be"); wres(15,"s21x")
        # S2.2
        wre(16,"s22dr","s22de"); wre(18,"s22cr","s22ce"); wre(20,"s22pr","s22pe"); wres(22,"s22x")
        # S2.3
        wre(23,"s23dr","s23de"); wres(25,"s23x"); wres(26,"s2x")
        # S3.1
        wre(27,"s31mr","s31me","EEF7EE"); wre(29,"s31cr","s31ce","EEF7EE")
        wre(31,"s31rr","s31re","EEF7EE"); wre(33,"s31sr","s31se","EEF7EE"); wres(35,"s31x")
        # S3.2
        wre(36,"s32nr","s32ne","EEF7EE"); wre(38,"s32or","s32oe","EEF7EE")
        wre(40,"s32dr","s32de","EEF7EE"); wres(42,"s32x"); wres(43,"s3x")
        # S4.1
        wre(44,"s41br","s41be","F5EEF8"); wre(46,"s41rr","s41re","F5EEF8")
        wre(48,"s41cr","s41ce","F5EEF8"); wre(50,"s41vr","s41ve","F5EEF8")
        wre(52,"s41pr","s41pe","F5EEF8"); wres(54,"s4x")
        # S5.1
        wre(55,"s51fr","s51fe","FDF0EE"); wre(57,"s51ar","s51ae","FDF0EE")
        wre(59,"s51pr","s51pe","FDF0EE"); wre(61,"s51ir","s51ie","FDF0EE"); wres(63,"s51x")
        # S5.2
        wre(64,"s52sr","s52se","FDF0EE"); wre(66,"s52cr","s52ce","FDF0EE")
        wre(68,"s52pr","s52pe","FDF0EE"); wre(70,"s52rr","s52re","FDF0EE")
        wres(72,"s52x"); wres(73,"s5x")
        # Final
        wres(74,"gx",8)
        _status(ws.cell(row,75), b.get("estado","—"))
        for col,key,sz in [(76,"obs",7),(77,"mejora",7)]:
            c=ws.cell(row,col,value=b.get(key,"—"))
            bg=row_color; c.fill=_fill(bg); c.font=Font(size=sz,italic=True,color="444444")
            c.alignment=wrap; c.border=_border()

        ws.row_dimensions[row].height = 30

    last_row = DATA_START + len(BOLTS) - 1

    # ── Merge celdas de Intent ────────────────────────────────────────────────
    run_s, run_id = DATA_START, BOLTS[0]["iid"]
    for i, b in enumerate(BOLTS[1:], 1):
        r = DATA_START + i
        if b["iid"] != run_id:
            if r-1 > run_s:
                ws.merge_cells(start_row=run_s,start_column=1,end_row=r-1,end_column=1)
                ws.cell(run_s,1).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            run_s,run_id = r, b["iid"]
    if last_row > run_s:
        ws.merge_cells(start_row=run_s,start_column=1,end_row=last_row,end_column=1)
        ws.cell(run_s,1).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

    # ── Fila resumen ──────────────────────────────────────────────────────────
    sr = last_row + 2
    si  = sum(1 for b in BOLTS if b["status"]=="Si")
    par = sum(1 for b in BOLTS if b["status"]=="Parcial")
    no  = sum(1 for b in BOLTS if b["status"]=="No")
    pct = round((si + 0.5*par) / len(BOLTS) * 100, 1)

    ws.merge_cells(f"A{sr}:C{sr}")
    c=ws.cell(sr,1,value="RESUMEN TOTAL"); c.font=Font(bold=True,size=9,color="FFFFFF")
    c.fill=_fill(NAVY); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=_border()

    summary_data = [
        (4,  f"Bolts totales: {len(BOLTS)}"),
        (5,  f"Completados: {si}"),
        (6,  f"En progreso: {par}"),
        (7,  f"Pendientes: {no}"),
        (8,  f"Avance: {pct} %"),
        (9,  "Cobertura backend: 100 %"),
        (10, "Cobertura frontend: 100 %"),
        (11, "SonarCloud: Quality Gate Aprobado"),
        (12, "Intents: 15 | Tests: 507"),
    ]
    for col, txt in summary_data:
        c = ws.cell(sr, col, value=txt)
        c.font = Font(bold=True, size=8)
        c.fill = _fill("EEF4FF")
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sr].height = 18

    # ── Formato condicional en columnas de resultado ──────────────────────────
    g_fill = PatternFill("solid", fgColor="C6EFCE")
    g_font = Font(bold=True, size=8, color="276221")
    r_fill = PatternFill("solid", fgColor="FFC7CE")
    r_font = Font(bold=True, size=8, color="9C0006")
    y_fill = PatternFill("solid", fgColor="FFEB9C")
    y_font = Font(bold=True, size=8, color="7D4B12")

    for rc in RESULT_COLS:
        rng = f"{cl(rc)}{DATA_START}:{cl(rc)}{last_row}"
        col_ref = f"{cl(rc)}{DATA_START}"
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Aprobado",{col_ref}))'], fill=g_fill, font=g_font))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Completado",{col_ref}))'], fill=g_fill, font=g_font))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Pendiente",{col_ref}))'], fill=r_fill, font=r_font))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("En progreso",{col_ref}))'], fill=y_fill, font=y_font))

    # ── Auto-filtro en fila 6 ─────────────────────────────────────────────────
    ws.auto_filter.ref = f"B6:{cl(8)}{last_row}"

    # ── Ancho de columnas ─────────────────────────────────────────────────────
    ws.column_dimensions[cl(1)].width  = 13   # Intent
    ws.column_dimensions[cl(2)].width  = 19   # ID del Bolt
    ws.column_dimensions[cl(3)].width  = 40   # Descripcion
    ws.column_dimensions[cl(4)].width  = 15   # Modelo IA
    ws.column_dimensions[cl(5)].width  = 10   # Fecha
    ws.column_dimensions[cl(6)].width  = 12   # Dominio
    ws.column_dimensions[cl(7)].width  = 10   # HU
    ws.column_dimensions[cl(8)].width  = 14   # Responsable

    for col in range(9, 78):
        if col in RESULT_COLS:
            ws.column_dimensions[cl(col)].width = 10
        elif col in (76, 77):
            ws.column_dimensions[cl(col)].width = 24
        elif col % 2 == 0:   # Evidencia (par)
            ws.column_dimensions[cl(col)].width = 26
        else:                 # Registro (impar)
            ws.column_dimensions[cl(col)].width = 7

    # ── Freeze panes y configuracion de impresion ─────────────────────────────
    ws.freeze_panes = "D7"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize   = ws.PAPERSIZE_A3
    ws.print_area = f"A1:{cl(MAX_COL)}{sr}"
    ws.oddHeader.center.text = "HearGuard AI — Matriz de Registro AI-DLC"
    ws.oddFooter.right.text  = "Pagina &P de &N"

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(OUT)
    if "Matriz de registro" in wb.sheetnames:
        idx = wb.sheetnames.index("Matriz de registro")
        del wb["Matriz de registro"]
    else:
        idx = 2
    ws = wb.create_sheet("Matriz de registro", idx)
    write_matriz(ws)
    wb.save(OUT)

    si  = sum(1 for b in BOLTS if b["status"]=="Si")
    par = sum(1 for b in BOLTS if b["status"]=="Parcial")
    no  = sum(1 for b in BOLTS if b["status"]=="No")
    pct = round((si+0.5*par)/len(BOLTS)*100,1)
    print(f"OK  Guardado: {OUT}")
    print(f"    Bolts: {len(BOLTS)} | Completados: {si} | En progreso: {par} | Pendientes: {no} | Avance: {pct}%")
    print(f"    Columnas AI-DLC: 77 | Intents: 15 | Mejoras: alternancia,CF,auto-filtro,impresion")

if __name__ == "__main__":
    main()
