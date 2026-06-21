#!/usr/bin/env python3
"""
Genera docs/matriz-registro-hearguard.xlsx — Matriz de registro HearGuard AI (UC).

Uso: python scripts/generar-matriz-registro.py
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

# ── Metadatos del proyecto (editar aquí si cambia el curso) ───────────────────
PROJECT = {
    "nombre": "HearGuard AI",
    "subtitulo": "Plataforma de salud auditiva preventiva con TDD/BDD + CRISP-DM",
    "autor": "Luis Francisco Terreros Hinojosa",
    "asesor": "Maglioni Arana Caparachín",
    "institucion": "Universidad Continental",
    "escuela": "Escuela Académico Profesional de Ingeniería de Sistemas e Informática",
    "curso": "Proyecto de titulación / Trabajo de investigación",
    "periodo": "2026-I",
    "codigo_curso": "[Código del curso — completar]",
    "email": "luisterreroshinojosa@gmail.com",
    "version_matriz": "3.0",
}

REPO_BASE = "https://github.com/hatWHITE-UwU/hearguard-ai"
REPO_COMMIT = f"{REPO_BASE}/commit"
SONAR_URL = "https://sonarcloud.io/project/overview?id=hatWHITE-UwU_hearguard-ai"
CI_URL = f"{REPO_BASE}/actions"
DOCS_TRAZA = f"{REPO_BASE}/blob/main/docs/matriz-trazabilidad.md"


@dataclass
class Bolt:
    intent_id: str
    intent_title: str
    bolt_id: str
    description: str
    model: str
    date: str
    domain: str
    methodology: str
    status: str
    evidence: str
    evidence_url: str | None = None
    trace: str = ""
    notes: str = ""


@dataclass
class RfGroup:
    rf_id: str
    nombre: str
    descripcion: str
    feature_bdd: str
    subrequisitos: int
    estado: str
    tests_principales: str
    intents: str


@dataclass
class RnfItem:
    rnf_id: str
    requisito: str
    verificacion: str
    estado: str
    evidencia: str


RF_GROUPS: list[RfGroup] = [
    RfGroup("RF-01", "Autenticación y sesión",
            "Registro, login, refresh JWT, logout, perfil y guards Angular.",
            "docs/features/autenticacion.feature", 17, "✅ Completo",
            "auth.test.js · security.test.js · auth.*.spec.ts", "INT-001, INT-002, INT-005"),
    RfGroup("RF-02", "Monitoreo de ruido",
            "Captura dB, clasificación de riesgo, historial y estadísticas hoy/semana.",
            "docs/features/monitoreo-ruido.feature", 15, "✅ Completo",
            "noise.test.js · noise-monitor.service.spec.ts", "INT-001, INT-002"),
    RfGroup("RF-03", "Prueba auditiva tonal",
            "Cuestionario 12 pasos (6 Hz × 2 oídos), scoring y evaluación complete/partial.",
            "docs/features/prueba-auditiva.feature", 13, "✅ Completo",
            "hearing-test.service.spec.ts · evaluation.test.js", "INT-001, INT-002"),
    RfGroup("RF-04", "Predicción de riesgo (IA)",
            "Random Forest: score 0-100, niveles, recomendaciones, integración Flask.",
            "docs/features/prediccion-riesgo-ia.feature", 15, "✅ Completo",
            "test_predictor.py · test_api.py · evaluation-ai.test.js", "INT-001, INT-006"),
    RfGroup("RF-05", "Resultados y recomendaciones",
            "Historial de evaluaciones, riskResult y visualización de resultados.",
            "docs/features/resultados-y-recomendaciones.feature", 6, "⚠️ E2E parcial",
            "evaluation.test.js · e2e/hearing-test.spec.ts", "INT-001, INT-002"),
    RfGroup("RF-06", "Dispositivos IoT (ESP32)",
            "Registro de dispositivos, X-Device-Key y POST /api/noise/iot.",
            "docs/features/dispositivos-iot.feature", 6, "✅ Completo",
            "device.test.js · noise.test.js (IoT)", "INT-001, INT-002, INT-005"),
]

RNF_ITEMS: list[RnfItem] = [
    RnfItem("RNF-01", "p95 respuesta API < 2 000 ms", "k6 threshold http_req_duration", "✅ Configurado",
            "tests/k6/load-test.js"),
    RnfItem("RNF-02", "Tasa de error < 5 % bajo carga", "k6 threshold http_req_failed", "✅ Configurado",
            "tests/k6/load-test.js"),
    RnfItem("RNF-03", "Cobertura backend 100 % (CI)", "Jest lcov + job backend ci.yml", "✅ Cumplido",
            "backend/coverage/lcov.info"),
    RnfItem("RNF-04", "Cobertura IA ≥ 60 %", "pytest --cov-fail-under=60", "✅ Cumplido",
            "ai-service/coverage.xml"),
    RnfItem("RNF-05", "ESLint sin errores (backend/frontend)", "npm run lint en CI", "✅ Cumplido",
            ".github/workflows/ci.yml"),
    RnfItem("RNF-06", "Protección NoSQL injection", "security.test.js", "✅ Cumplido",
            "backend/tests/security.test.js"),
    RnfItem("RNF-07", "JWT HS256 / rechazo alg:none", "security.test.js", "✅ Cumplido",
            "backend/tests/security.test.js"),
    RnfItem("RNF-08", "Rate limiting API", "server.js apiLimiter", "✅ Cumplido",
            "backend/tests/security.test.js"),
    RnfItem("RNF-09", "SonarCloud Quality Gate OK", "job sonarcloud + SONAR_TOKEN", "✅ Cumplido",
            SONAR_URL),
    RnfItem("RNF-10", "Conventional Commits (Husky)", ".husky/commit-msg", "✅ Cumplido",
            ".husky/"),
]

# Bolts (misma base ampliada; ver repo para historial completo)
BOLTS: list[Bolt] = [
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-001",
         "Configurar Jest + Supertest + cobertura lcov (umbral CI 100 %)",
         "Cursor Composer", "03/2026", "Backend", "TDD", "Sí", "backend/package.json", None, "RF-01–06", ""),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-002",
         "Suite integración: auth, noise, evaluation, device, middleware",
         "Cursor Composer", "03/2026", "Backend", "TDD", "Sí", "backend/tests/ (207 tests)", None, "RF-01–06", ""),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-003",
         "Tests seguridad: JWT, IDOR, NoSQL, rutas protegidas (22 casos)",
         "Cursor Composer", "03/2026", "Seguridad", "TDD", "Sí", "security.test.js", None, "RF-01, RF-06", ""),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-004",
         "pytest + coverage XML en ai-service (30 tests)",
         "Cursor Composer", "03/2026", "IA", "TDD", "Sí", "ai-service/tests/", None, "RF-04", ""),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-005",
         "Vitest Angular: core + features services (107 tests)",
         "Cursor Composer", "03/2026", "Frontend", "TDD", "Sí", "frontend/**/*.spec.ts", None, "RF-01, RF-03", ""),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-006",
         "flutter_test móvil (42 tests)",
         "Cursor Composer", "03/2026", "Móvil", "TDD", "Sí", "flutter_app/test/", None, "—", ""),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-007",
         "Playwright E2E smoke + auth + hearing-test (36 tests)",
         "Cursor Composer", "03/2026", "E2E", "TDD", "Sí", "e2e/tests/", None, "RF-05", "Vercel preview"),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-008",
         "Plan de pruebas IEEE 829 / ISO 29119",
         "Cursor Composer", "03/2026", "Docs", "TDD", "Sí", "docs/plan-de-pruebas.md", None, "—", "1169 líneas"),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-009",
         "coverage-extra + noise.service + database tests",
         "Cursor Composer", "21/05/2026", "Backend", "TDD", "Sí", "Commit 8ebc768", f"{REPO_COMMIT}/8ebc768", "—", "100 % líneas"),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-010",
         "evaluation-ai.test.js — flujo IA exitoso",
         "Cursor Composer", "22/05/2026", "Backend", "TDD", "Sí", "Commit f8daa25", f"{REPO_COMMIT}/f8daa25", "RF-04", ""),
    Bolt("INT-001", "Framework de testing (TDD)", "INT-001-BOLT-011",
         "Jest --runInBand para estabilidad MongoDB en cobertura",
         "Equipo", "22/05/2026", "Backend", "TDD", "Sí", "README § Tests", None, "—", ""),
    Bolt("INT-002", "BDD — Gherkin y trazabilidad", "INT-002-BOLT-001",
         "6 archivos .feature (auth, ruido, auditiva, IA, IoT, resultados)",
         "Cursor Composer", "03/2026", "BDD", "BDD", "Sí", "docs/features/", None, "RF-01–06", ""),
    Bolt("INT-002", "BDD — Gherkin y trazabilidad", "INT-002-BOLT-002",
         "Matriz trazabilidad RF ↔ BDD ↔ test (60 RF, 10 RNF)",
         "Cursor Composer", "03/2026", "Docs", "BDD", "Sí", "docs/matriz-trazabilidad.md", DOCS_TRAZA, "—", ""),
    Bolt("INT-002", "BDD — Gherkin y trazabilidad", "INT-002-BOLT-003",
         "Metodología TDD+BDD documentada con referencias APA",
         "Cursor Composer", "05/2026", "Docs", "BDD", "Sí", "docs/metodologia.md", None, "—", ""),
    Bolt("INT-002", "BDD — Gherkin y trazabilidad", "INT-002-BOLT-004",
         "Ejecutar .feature con Cucumber en CI",
         "—", "—", "BDD", "BDD", "No", "Futuro", None, "—", "Validación vía Jest/pytest"),
    Bolt("INT-003", "CI/CD — Validaciones automáticas", "INT-003-BOLT-001",
         "Jobs: backend, ai-service, frontend, e2e, flutter, sonarcloud",
         "Cursor Composer", "03–05/2026", "CI/CD", "CI", "Sí", "ci.yml", CI_URL, "RNF-09", ""),
    Bolt("INT-003", "CI/CD — Validaciones automáticas", "INT-003-BOLT-002",
         "Artefactos cobertura + fix-sonar-coverage-paths.js",
         "Cursor Composer", "21/05/2026", "CI/CD", "CI", "Sí", "scripts/fix-sonar-coverage-paths.js", None, "RNF-03", ""),
    Bolt("INT-003", "CI/CD — Validaciones automáticas", "INT-003-BOLT-003",
         "deploy.yml → Render + Vercel",
         "Cursor Composer", "03/2026", "Deploy", "CI", "Sí", "deploy.yml", None, "—", ""),
    Bolt("INT-003", "CI/CD — Validaciones automáticas", "INT-003-BOLT-004",
         "Badge CI verde en README",
         "GitHub Actions", "05/2026", "CI/CD", "CI", "Sí", "README badge", CI_URL, "—", ""),
    Bolt("INT-003", "CI/CD — Validaciones automáticas", "INT-003-BOLT-005",
         "Optimizar duración pipeline (métricas SLA)",
         "—", "—", "CI/CD", "CI", "Parcial", "Jobs paralelos OK", None, "—", "Sin benchmark formal"),
    Bolt("INT-004", "Calidad — SonarCloud", "INT-004-BOLT-001",
         "sonar-project.properties + exclusiones",
         "Cursor Composer", "05/2026", "Calidad", "SAST", "Sí", "sonar-project.properties", SONAR_URL, "RNF-09", ""),
    Bolt("INT-004", "Calidad — SonarCloud", "INT-004-BOLT-002",
         "Ratings Security / Reliability / Maintainability = A",
         "Cursor Composer", "21/05/2026", "Calidad", "SAST", "Sí", "Commit 28925d5", f"{REPO_COMMIT}/28925d5", "—", "0 issues"),
    Bolt("INT-004", "Calidad — SonarCloud", "INT-004-BOLT-003",
         "Cerrar S2699 y code smells (37+)",
         "Cursor Composer", "20–21/05/2026", "Calidad", "SAST", "Sí", "Commits ddde986, 319e9bc", f"{REPO_COMMIT}/319e9bc", "—", ""),
    Bolt("INT-004", "Calidad — SonarCloud", "INT-004-BOLT-004",
         "Duplicación 0 % + cobertura 100 % Sonar",
         "Cursor Composer", "22/05/2026", "Calidad", "SAST", "Sí", "Commit 8ebc768", f"{REPO_COMMIT}/8ebc768", "RNF-03", ""),
    Bolt("INT-005", "Seguridad", "INT-005-BOLT-001",
         "S5147 NoSQL injection + $eq en Device/Noise",
         "Cursor Composer", "19–20/05/2026", "Seguridad", "SAST", "Sí", "25c1603, 1998062", f"{REPO_COMMIT}/25c1603", "RNF-06", ""),
    Bolt("INT-005", "Seguridad", "INT-005-BOLT-002",
         "S2068 + PRNG seguro en E2E helpers",
         "Cursor Composer", "20/05/2026", "Seguridad", "SAST", "Sí", "fb4119c, 8c5d890", f"{REPO_COMMIT}/fb4119c", "—", "Regresión C→A"),
    Bolt("INT-005", "Seguridad", "INT-005-BOLT-003",
         "npm audit periódico",
         "Equipo", "—", "Seguridad", "SAST", "Parcial", "Manual", None, "—", "Sin job tipo bundler-audit"),
    Bolt("INT-006", "Modelo IA — CRISP-DM", "INT-006-BOLT-001",
         "Fases 1-2: negocio y variables",
         "Cursor Composer", "03/2026", "IA", "CRISP-DM", "Sí", "README · metodologia.md", None, "RF-04", ""),
    Bolt("INT-006", "Modelo IA — CRISP-DM", "INT-006-BOLT-002",
         "Fase 3: features.py + constants.py",
         "Cursor Composer", "03/2026", "IA", "CRISP-DM", "Sí", "ai-service/model/", None, "RF-04", "8 features"),
    Bolt("INT-006", "Modelo IA — CRISP-DM", "INT-006-BOLT-003",
         "Fase 4: trainer Random Forest SEED=42",
         "Cursor Composer", "03/2026", "IA", "CRISP-DM", "Sí", "trainer.py", None, "RF-04", ""),
    Bolt("INT-006", "Modelo IA — CRISP-DM", "INT-006-BOLT-004",
         "Fase 5: R²≥0.80 + perfiles bajo/alto",
         "Cursor Composer", "03/2026", "IA", "CRISP-DM", "Sí", "test_predictor.py", None, "RF-04", ""),
    Bolt("INT-006", "Modelo IA — CRISP-DM", "INT-006-BOLT-005",
         "Fase 6: Flask + Render + ai.service.js",
         "Cursor Composer", "03/2026", "IA", "CRISP-DM", "Sí", "app.py · render.yaml", None, "RF-04", ""),
    Bolt("INT-006", "Modelo IA — CRISP-DM", "INT-006-BOLT-006",
         "Reentrenamiento en CI (job ai-service)",
         "Cursor Composer", "05/2026", "IA", "CRISP-DM", "Sí", "ci.yml", CI_URL, "—", ""),
    Bolt("INT-007", "Rendimiento (RNF)", "INT-007-BOLT-001",
         "k6: smoke, load, spike + umbrales p95/error",
         "Cursor Composer", "03/2026", "Rendimiento", "RNF", "Sí", "tests/k6/load-test.js", None, "RNF-01, RNF-02", ""),
    Bolt("INT-007", "Rendimiento (RNF)", "INT-007-BOLT-002",
         "Ejecutar k6 en Render y adjuntar reporte al informe",
         "—", "—", "Rendimiento", "RNF", "No", "Pendiente entrega", None, "RNF-01", "BASE_URL=..."),
    Bolt("INT-007", "Rendimiento (RNF)", "INT-007-BOLT-003",
         "Lighthouse en frontend Vercel",
         "—", "—", "Frontend", "RNF", "No", "Chrome DevTools", None, "—", ""),
    Bolt("INT-008", "Documentación y entregables", "INT-008-BOLT-001",
         "articulo.md + README + api-spec",
         "Cursor Composer", "05/2026", "Docs", "—", "Sí", "docs/articulo.md", None, "—", ""),
    Bolt("INT-008", "Documentación y entregables", "INT-008-BOLT-002",
         "Matriz de registro Excel v3 (este archivo)",
         "Cursor Composer", datetime.now().strftime("%d/%m/%Y"), "Docs", "—", "Sí",
         "docs/matriz-registro-hearguard.xlsx", None, "—", f"v{PROJECT['version_matriz']}"),
    Bolt("INT-008", "Documentación y entregables", "INT-008-BOLT-003",
         "complejidad-ciclomatica.md (McCabe)",
         "Cursor Composer", "03/2026", "Docs", "—", "Sí", "docs/complejidad-ciclomatica.md", None, "—", ""),
    Bolt("INT-008", "Documentación y entregables", "INT-008-BOLT-004",
         "Docker Compose reproducible",
         "Cursor Composer", "03/2026", "DevOps", "—", "Sí", "docker-compose.yml", None, "—", ""),
]

INTENT_COLORS = {
    "INT-001": "E2EFDA", "INT-002": "DDEBF7", "INT-003": "FFF2CC",
    "INT-004": "FCE4D6", "INT-005": "E7E6E6", "INT-006": "E4DFEC",
    "INT-007": "D9E1F2", "INT-008": "F8CBAD",
}
STATUS_FILL = {"Sí": "C6EFCE", "No": "FFC7CE", "Parcial": "FFEB9C"}
BRAND = "1F4E79"
ACCENT = "2E75B6"


def _border() -> Border:
    s = Side(style="thin", color="B4B4B4")
    return Border(left=s, right=s, top=s, bottom=s)


def _hyperlink(cell, url: str, label: str | None = None) -> None:
    if url and url.startswith("http"):
        cell.value = label or "Abrir enlace"
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url)
        cell.font = Font(color="0563C1", underline="single")


def _style_header_row(ws, row: int, cols: int, titles: list[str]) -> None:
    for c, title in enumerate(titles, 1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()


def _write_portada(wb: Workbook) -> None:
    ws = wb.create_sheet("Portada", 0)
    ws.sheet_view.showGridLines = False
    gen = datetime.now().strftime("%d/%m/%Y %H:%M")

    blocks = [
        (2, PROJECT["institucion"], 18, True),
        (3, PROJECT["escuela"], 12, False),
        (5, PROJECT["nombre"], 22, True),
        (6, PROJECT["subtitulo"], 12, False),
        (8, "MATRIZ DE REGISTRO DE ACTIVIDADES", 16, True),
        (9, f"Versión {PROJECT['version_matriz']} · Generada: {gen}", 10, False),
        (11, "Datos del estudiante", 12, True),
        (12, f"Autor: {PROJECT['autor']}", 11, False),
        (13, f"Correo: {PROJECT['email']}", 11, False),
        (14, f"Asesor: {PROJECT['asesor']}", 11, False),
        (15, f"Curso / proyecto: {PROJECT['curso']}", 11, False),
        (16, f"Periodo académico: {PROJECT['periodo']}", 11, False),
        (17, f"Código de curso: {PROJECT['codigo_curso']}", 11, False),
        (19, "Metodologías aplicadas", 12, True),
        (20, "Principal: Test-Driven Development (TDD) + Behavior-Driven Development (BDD)", 11, False),
        (21, "Complementaria: CRISP-DM (modelo predictivo Random Forest)", 11, False),
        (23, "Enlaces del proyecto", 12, True),
    ]
    for row, text, size, bold in blocks:
        ws.merge_cells(f"A{row}:H{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = Font(bold=bold, size=size, color=BRAND if bold else "333333")
        c.alignment = Alignment(horizontal="left", vertical="center")

    links = [
        ("Repositorio GitHub", REPO_BASE),
        ("SonarCloud (calidad)", SONAR_URL),
        ("GitHub Actions (CI)", CI_URL),
        ("Matriz de trazabilidad RF/RNF", DOCS_TRAZA),
    ]
    r = 24
    for label, url in links:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=11)
        lc = ws.cell(row=r, column=2, value=url)
        _hyperlink(lc, url, url)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"].value = (
        "Navegación: use las pestañas inferiores — «Dashboard», «Matriz de registro», "
        "«Requisitos funcionales», «Requisitos no funcionales»."
    )
    ws[f"A{r}"].font = Font(italic=True, size=10, color="666666")
    ws[f"A{r}"].alignment = Alignment(wrap_text=True)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["B"].width = 55
    ws.row_dimensions[5].height = 36


def _write_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Dashboard — HearGuard AI"
    ws["A1"].font = Font(bold=True, size=14, color=BRAND)

    intents: dict[str, dict] = {}
    for b in BOLTS:
        if b.intent_id not in intents:
            intents[b.intent_id] = {"title": b.intent_title, "t": 0, "Sí": 0, "No": 0, "Parcial": 0}
        intents[b.intent_id]["t"] += 1
        intents[b.intent_id][b.status] += 1

    headers = ["Intent", "Objetivo", "Bolts", "Sí", "No", "Parcial", "% Avance"]
    _style_header_row(ws, 3, 7, headers)

    row = 4
    chart_labels = []
    chart_vals = []
    for iid in sorted(intents.keys()):
        d = intents[iid]
        pct = round((d["Sí"] + 0.5 * d["Parcial"]) / d["t"] * 100, 1)
        ws.cell(row=row, column=1, value=iid)
        ws.cell(row=row, column=2, value=d["title"])
        ws.cell(row=row, column=3, value=d["t"])
        ws.cell(row=row, column=4, value=d["Sí"])
        ws.cell(row=row, column=5, value=d["No"])
        ws.cell(row=row, column=6, value=d["Parcial"])
        pc = ws.cell(row=row, column=7, value=pct / 100)
        pc.number_format = "0.0%"
        if pct >= 90:
            pc.fill = PatternFill("solid", fgColor="C6EFCE")
        elif pct >= 70:
            pc.fill = PatternFill("solid", fgColor="FFEB9C")
        else:
            pc.fill = PatternFill("solid", fgColor="FFC7CE")
        chart_labels.append(iid)
        chart_vals.append(pct)
        row += 1

    total = len(BOLTS)
    si = sum(1 for b in BOLTS if b.status == "Sí")
    no = sum(1 for b in BOLTS if b.status == "No")
    par = sum(1 for b in BOLTS if b.status == "Parcial")
    g_pct = round((si + 0.5 * par) / total * 100, 1)

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for c, v in enumerate([total, si, no, par, g_pct / 100], 3):
        cell = ws.cell(row=row, column=c, value=v)
        if c == 7:
            cell.number_format = "0.0%"
        cell.font = Font(bold=True)

    # KPIs
    krow = row + 3
    kpis = [
        ("Autor", PROJECT["autor"]),
        ("Requisitos funcionales", f"{sum(g.subrequisitos for g in RF_GROUPS)} sub-RF en 6 módulos"),
        ("Requisitos no funcionales", f"{len(RNF_ITEMS)} RNF"),
        ("Pruebas automatizadas", "422 + 3 escenarios k6"),
        ("Cobertura SonarCloud", "100 % · Duplicación 0 %"),
        ("Ratings Sonar", "Security A · Reliability A · Maintainability A"),
        ("Bolts completados", f"{si} / {total} ({g_pct}%)"),
        ("Pendientes clave", "k6 prod · Lighthouse · Cucumber CI"),
    ]
    ws.cell(row=krow, column=1, value="Indicadores clave").font = Font(bold=True, size=12, color=BRAND)
    krow += 1
    for label, val in kpis:
        ws.cell(row=krow, column=1, value=label).font = Font(bold=True)
        ws.cell(row=krow, column=2, value=val)
        krow += 1

    # Gráfico barras % avance por Intent
    chart = BarChart()
    chart.type = "col"
    chart.title = "% Avance por Intent"
    chart.y_axis.title = "Porcentaje"
    chart.x_axis.title = "Intent"
    data = Reference(ws, min_col=7, min_row=3, max_row=3 + len(intents))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(intents))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 22
    ws.add_chart(chart, f"I{3}")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    for c in "CDEFG":
        ws.column_dimensions[c].width = 10


def _merge_intent_cells(ws, col: int, start_row: int, end_row: int, bolts: list[Bolt]) -> None:
    """Fusiona celdas de Intent cuando el mismo intent_id se repite."""
    if not bolts:
        return
    run_start = start_row
    current = bolts[0].intent_id
    for i, bolt in enumerate(bolts[1:], start=1):
        r = start_row + i
        if bolt.intent_id != current:
            if r - 1 > run_start:
                ws.merge_cells(
                    start_row=run_start, start_column=col,
                    end_row=r - 1, end_column=col,
                )
                merged = ws.cell(run_start, col)
                merged.alignment = Alignment(vertical="center", wrap_text=True)
            run_start = r
            current = bolt.intent_id
    if end_row > run_start:
        ws.merge_cells(start_row=run_start, start_column=col, end_row=end_row, end_column=col)
        ws.cell(run_start, col).alignment = Alignment(vertical="center", wrap_text=True)


def _write_matriz(wb: Workbook) -> None:
    ws = wb.create_sheet("Matriz de registro")
    headers = [
        "Intent", "ID del Bolt", "Descripción del Bolt",
        "Modelo IA", "Fecha", "Dominio", "Metodología",
        "RF / RNF", "Registro", "Evidencia", "Enlace", "Observaciones",
    ]

    ws.merge_cells("A1:L1")
    ws["A1"] = f"{PROJECT['nombre']} — Matriz de registro (Bolts)"
    ws["A1"].font = Font(bold=True, size=14, color=BRAND)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        f"{PROJECT['autor']} · {PROJECT['institucion']} · "
        f"Metodología: TDD+BDD + CRISP-DM"
    )
    ws["A2"].font = Font(size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Subsecciones estilo curso
    sub_row = 4
    ws.merge_cells(f"D{sub_row}:G{sub_row}")
    ws[f"D{sub_row}"] = "1. Identificación"
    ws[f"D{sub_row}"].font = Font(bold=True, size=9, color=ACCENT)
    ws[f"D{sub_row}"].alignment = Alignment(horizontal="center")
    ws[f"D{sub_row}"].fill = PatternFill("solid", fgColor="D6E4F0")

    ws.merge_cells(f"H{sub_row}:L{sub_row}")
    ws[f"H{sub_row}"] = "2. Registro y evidencia"
    ws[f"H{sub_row}"].font = Font(bold=True, size=9, color=ACCENT)
    ws[f"H{sub_row}"].alignment = Alignment(horizontal="center")
    ws[f"H{sub_row}"].fill = PatternFill("solid", fgColor="D6E4F0")

    header_row = 5
    _style_header_row(ws, header_row, len(headers), headers)
    data_start = header_row + 1
    wrap = Alignment(wrap_text=True, vertical="top")

    for i, bolt in enumerate(BOLTS):
        r = data_start + i
        intent_text = f"{bolt.intent_id}\n{bolt.intent_title}"
        vals = [
            intent_text, bolt.bolt_id, bolt.description,
            bolt.model, bolt.date, bolt.domain, bolt.methodology,
            bolt.trace, bolt.status, bolt.evidence, "", bolt.notes,
        ]
        color = INTENT_COLORS.get(bolt.intent_id, "FFFFFF")
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = wrap
            cell.border = _border()
            if col == 1:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, size=9)
            if col == 9:
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL.get(bolt.status, "FFFFFF"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 11 and bolt.evidence_url:
                _hyperlink(cell, bolt.evidence_url)

        ws.row_dimensions[r].height = 36

    last_row = data_start + len(BOLTS) - 1
    _merge_intent_cells(ws, 1, data_start, last_row, BOLTS)

    widths = [22, 17, 50, 18, 11, 14, 12, 12, 10, 30, 14, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = f"A{header_row}:L{last_row}"

    dv = DataValidation(type="list", formula1='"Sí,No,Parcial"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"I{data_start}:I{last_row}")


def _write_rf_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Requisitos funcionales")
    ws["A1"] = "Requisitos funcionales (RF) — Resumen HearGuard AI"
    ws["A1"].font = Font(bold=True, size=13, color=BRAND)
    ws.merge_cells("A2:I2")
    ws["A2"] = (
        f"Detalle completo: docs/matriz-trazabilidad.md · "
        f"Total: {sum(g.subrequisitos for g in RF_GROUPS)} sub-requisitos trazados a tests"
    )
    ws["A2"].font = Font(italic=True, size=10)

    headers = [
        "RF", "Módulo", "Descripción", "Feature BDD",
        "N.° sub-RF", "Estado", "Tests principales", "Intents relacionados",
    ]
    _style_header_row(ws, 4, len(headers), headers)

    for i, g in enumerate(RF_GROUPS, 5):
        row = [
            g.rf_id, g.nombre, g.descripcion, g.feature_bdd,
            g.subrequisitos, g.estado, g.tests_principales, g.intents,
        ]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _border()
            if c == 6 and "✅" in str(val):
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif c == 6 and "⚠️" in str(val):
                cell.fill = PatternFill("solid", fgColor="FFEB9C")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 22
    ws.freeze_panes = "A5"


def _write_rnf_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Requisitos no funcionales")
    ws["A1"] = "Requisitos no funcionales (RNF) — HearGuard AI"
    ws["A1"].font = Font(bold=True, size=13, color=BRAND)

    headers = ["RNF", "Requisito", "Verificación / herramienta", "Estado", "Evidencia"]
    _style_header_row(ws, 3, len(headers), headers)

    for i, item in enumerate(RNF_ITEMS, 4):
        for c, val in enumerate(
            [item.rnf_id, item.requisito, item.verificacion, item.estado, item.evidencia], 1
        ):
            cell = ws.cell(row=i, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _border()
            if c == 4:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
            if c == 5 and str(val).startswith("http"):
                _hyperlink(cell, val, "Ver enlace")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 42
    ws.freeze_panes = "A4"


def _write_metricas(wb: Workbook) -> None:
    ws = wb.create_sheet("Métricas pruebas")
    ws["A1"] = "Métricas de pruebas por capa"
    ws["A1"].font = Font(bold=True, size=12, color=BRAND)

    rows = [
        ("Capa", "Framework", "Ubicación", "Casos", "Comando verificación"),
        ("Backend API", "Jest + Supertest", "backend/tests/", 230,
         "cd backend && npm test -- --runInBand"),
        ("Seguridad API", "Jest", "backend/tests/security.test.js", 22, "(incluido en backend)"),
        ("Servicio IA", "pytest", "ai-service/tests/", 30, "pytest tests/ -v --cov"),
        ("Frontend web", "Vitest", "frontend/src/app/**/*.spec.ts", 107, "npm run test:ci"),
        ("App móvil", "flutter_test", "flutter_app/test/", 42, "flutter test"),
        ("E2E web", "Playwright", "e2e/tests/", 36, "npx playwright test --project=chromium"),
        ("Rendimiento", "Grafana k6", "tests/k6/", "3 escenarios", "k6 run tests/k6/load-test.js"),
        ("BDD Gherkin", "Cucumber.js", "bdd/ + docs/features/", 85, "cd bdd && npm test"),
        ("TOTAL automatizado", "—", "—", 530, ""),
    ]
    for r, row in enumerate(rows, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = _border()
            if r == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=BRAND)
            if r == len(rows) + 2:
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["E"].width = 52


def _write_iso25010_matrix(wb: Workbook) -> None:
    """Matriz de doble entrada: Módulos × Características ISO/IEC 25010:2011."""
    ws = wb.create_sheet("ISO 25010 — Calidad")
    ws.sheet_view.showGridLines = False

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "Matriz de Calidad — HearGuard AI v1.0 × ISO/IEC 25010:2011"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=BRAND)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:J2")
    sub = ws["A2"]
    sub.value = (
        "● Cobertura principal   ○ Cobertura secundaria   — No aplica   "
        "| Verde = Principal · Azul = Secundaria · Gris = N/A"
    )
    sub.font = Font(italic=True, size=9, color="444444")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # ── Encabezados de columnas ───────────────────────────────────────────────
    ISO_CHARS = [
        "Adec.\nFuncional",
        "Efic.\nRendimiento",
        "Compa-\ntibilidad",
        "Usabi-\nlidad",
        "Fiabi-\nlidad",
        "Segu-\nridad",
        "Manteni-\nbilidad",
        "Porta-\nbilidad",
    ]

    ws.cell(row=3, column=1, value="ID").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=3, column=1).fill = PatternFill("solid", fgColor=BRAND)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=1).border = _border()

    ws.cell(row=3, column=2, value="Módulo / Componente").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=3, column=2).fill = PatternFill("solid", fgColor=BRAND)
    ws.cell(row=3, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=2).border = _border()

    for ci, label in enumerate(ISO_CHARS, 3):
        cell = ws.cell(row=3, column=ci, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[3].height = 36

    # ── Datos de la matriz ────────────────────────────────────────────────────
    FILL_P = PatternFill("solid", fgColor="C6EFCE")   # verde  — cobertura principal
    FILL_S = PatternFill("solid", fgColor="DDEEFF")   # azul   — cobertura secundaria
    FILL_N = PatternFill("solid", fgColor="F2F2F2")   # gris   — no aplica
    FONT_P = Font(bold=True, size=11, color="276221")
    FONT_S = Font(size=11, color="0B4F6C")
    FONT_N = Font(size=11, color="AAAAAA")

    MODULES = [
        ("RF-01", "Autenticación y sesión",            "●","○","○","○","●","●","●","○"),
        ("RF-02", "Monitoreo de ruido en tiempo real", "●","●","●","●","○","○","●","○"),
        ("RF-03", "Evaluación auditiva (cuestionario)","●","○","○","●","○","○","●","○"),
        ("RF-04", "Predicción de riesgo (IA Flask)",   "●","●","○","○","●","○","●","○"),
        ("RF-05", "Resultados y recomendaciones",      "●","○","○","●","○","●","●","○"),
        ("RF-06", "Dispositivos IoT (ESP32)",          "●","●","●","○","●","●","●","●"),
        ("Web",   "Frontend — Angular 21",             "○","●","●","●","○","○","●","●"),
        ("Móvil", "App Móvil — Flutter 3",             "○","●","●","●","○","○","●","●"),
        ("API",   "Backend — Node.js 20 / Express 5",  "○","●","●","○","●","●","●","●"),
        ("IA",    "Microservicio IA — Flask + RF",     "○","●","●","○","●","○","●","●"),
        ("BD",    "Base de datos — MongoDB Atlas",     "○","●","●","○","●","●","●","●"),
        ("IoT",   "Firmware — ESP32 + KY-037",         "○","●","●","—","●","●","●","●"),
        ("CI/CD", "Pipeline — GitHub Actions 10 jobs", "○","○","○","—","●","○","●","●"),
        ("SC",    "Análisis estático — SonarCloud",    "○","○","—","—","●","●","●","○"),
        ("Tests", "Suite 530 casos — TDD/BDD/E2E",     "●","○","○","○","●","●","●","○"),
        ("OWASP", "Seguridad — OWASP Top 10",          "○","—","—","—","●","●","●","○"),
        ("Infra", "Contenedores — Docker Compose",     "—","○","●","—","●","○","●","●"),
    ]

    fill_map = {"●": FILL_P, "○": FILL_S, "—": FILL_N}
    font_map = {"●": FONT_P, "○": FONT_S, "—": FONT_N}

    # alternating row bg for readability
    ALT_BG = PatternFill("solid", fgColor="F7F9FC")

    for ri, row_data in enumerate(MODULES, 4):
        mod_id, mod_name, *vals = row_data
        bg = ALT_BG if ri % 2 == 0 else None

        # ID cell
        c1 = ws.cell(row=ri, column=1, value=mod_id)
        c1.font = Font(bold=True, size=9)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = _border()
        if bg:
            c1.fill = bg

        # Name cell
        c2 = ws.cell(row=ri, column=2, value=mod_name)
        c2.font = Font(size=9)
        c2.alignment = Alignment(vertical="center", wrap_text=True)
        c2.border = _border()
        if bg:
            c2.fill = bg

        # Value cells
        for ci, sym in enumerate(vals, 3):
            cell = ws.cell(row=ri, column=ci, value=sym)
            cell.fill = fill_map.get(sym, FILL_N)
            cell.font = font_map.get(sym, FONT_N)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
        ws.row_dimensions[ri].height = 18

    # ── Fila de totales de cobertura principal ────────────────────────────────
    TOTAL_ROW = len(MODULES) + 4
    ws.merge_cells(f"A{TOTAL_ROW}:B{TOTAL_ROW}")
    tc = ws.cell(row=TOTAL_ROW, column=1, value="Módulos con cobertura principal (●)")
    tc.font = Font(bold=True, size=9, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor=BRAND)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = _border()
    ws.cell(row=TOTAL_ROW, column=2).border = _border()

    for ci, col_vals in enumerate(range(8), 3):
        count = sum(
            1 for row_data in MODULES
            if row_data[2 + (ci - 3)] == "●"
        )
        cell = ws.cell(row=TOTAL_ROW, column=ci, value=count)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[TOTAL_ROW].height = 20

    # ── Leyenda ───────────────────────────────────────────────────────────────
    LEG_ROW = TOTAL_ROW + 2
    ws.cell(row=LEG_ROW, column=1, value="Leyenda:").font = Font(bold=True, size=9)
    leyenda = [
        ("●", FILL_P, FONT_P, "Cobertura principal — el módulo contribuye directamente a esta característica"),
        ("○", FILL_S, FONT_S, "Cobertura secundaria — contribución parcial o indirecta"),
        ("—", FILL_N, FONT_N, "No aplica — la característica no es relevante para este módulo"),
    ]
    for i, (sym, fill, font, desc) in enumerate(leyenda, LEG_ROW + 1):
        sc = ws.cell(row=i, column=1, value=sym)
        sc.fill = fill
        sc.font = font
        sc.alignment = Alignment(horizontal="center")
        sc.border = _border()
        dc = ws.cell(row=i, column=2, value=desc)
        dc.font = Font(size=9)
        ws.merge_cells(f"B{i}:J{i}")

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 34
    for col_letter in [get_column_letter(c) for c in range(3, 11)]:
        ws.column_dimensions[col_letter].width = 11
    ws.freeze_panes = "C4"


def _write_instrucciones(wb: Workbook) -> None:
    ws = wb.create_sheet("Instrucciones")
    text = f"""
GUÍA DE USO — Matriz de registro HearGuard AI v{PROJECT['version_matriz']}

PESTAÑAS
  • Portada — Datos del autor ({PROJECT['autor']}) e institución.
  • Dashboard — Avance por Intent, KPIs y gráfico de barras.
  • Matriz de registro — Registro detallado de Bolts (como en el curso).
  • Requisitos funcionales — Resumen RF-01 a RF-06.
  • Requisitos no funcionales — RNF-01 a RNF-10.
  • Métricas pruebas — 530 tests + comandos.
  • ISO 25010 — Calidad — Matriz de doble entrada Módulos × Características ISO/IEC 25010:2011.
  • Instrucciones — Este texto.

COLUMNAS MATRIZ
  1. Identificación: Modelo IA, Fecha, Dominio, Metodología.
  2. Registro: RF/RNF vinculado, Sí/No/Parcial, Evidencia, Enlace GitHub/Sonar.

REGENERAR
  python scripts/generar-matriz-registro.py

PERSONALIZAR
  Edite PROJECT en scripts/generar-matriz-registro.py (asesor, código de curso).

EQUIVALENCIA CURSO (Rails → HearGuard)
  RSpec / FactoryBot  →  Jest, pytest, Vitest, flutter_test
  Brakeman          →  SonarCloud + security.test.js
  Bundler Audit     →  npm audit (manual)
  Cucumber          →  docs/features/*.feature (pendiente CI)
"""
    for i, line in enumerate(text.strip().split("\n"), 1):
        ws.cell(row=i, column=1, value=line)
        if line.startswith("GUÍA") or line.startswith("PESTAÑAS"):
            ws.cell(row=i, column=1).font = Font(bold=True, size=11, color=BRAND)
    ws.column_dimensions["A"].width = 95


def _write_resumen_intent(wb: Workbook) -> None:
    """Tabla compacta duplicada para impresión."""
    ws = wb.create_sheet("Resumen Intents")
    _style_header_row(ws, 1, 7, ["Intent", "Título", "Bolts", "Sí", "No", "Parcial", "%"])
    intents: dict[str, dict] = {}
    for b in BOLTS:
        if b.intent_id not in intents:
            intents[b.intent_id] = {"title": b.intent_title, "t": 0, "Sí": 0, "No": 0, "Parcial": 0}
        intents[b.intent_id]["t"] += 1
        intents[b.intent_id][b.status] += 1
    row = 2
    for iid in sorted(intents.keys()):
        d = intents[iid]
        pct = round((d["Sí"] + 0.5 * d["Parcial"]) / d["t"] * 100, 1)
        ws.cell(row=row, column=1, value=iid)
        ws.cell(row=row, column=2, value=d["title"])
        ws.cell(row=row, column=3, value=d["t"])
        ws.cell(row=row, column=4, value=d["Sí"])
        ws.cell(row=row, column=5, value=d["No"])
        ws.cell(row=row, column=6, value=d["Parcial"])
        ws.cell(row=row, column=7, value=f"{pct}%")
        row += 1


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    out = root / "docs" / "matriz-registro-hearguard.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    _write_portada(wb)
    _write_dashboard(wb)
    _write_matriz(wb)
    _write_rf_sheet(wb)
    _write_rnf_sheet(wb)
    _write_metricas(wb)
    _write_iso25010_matrix(wb)
    _write_resumen_intent(wb)
    _write_instrucciones(wb)

    wb.active = wb["Portada"]
    wb.save(out)

    total = len(BOLTS)
    si = sum(1 for b in BOLTS if b.status == "Sí")
    no = sum(1 for b in BOLTS if b.status == "No")
    par = sum(1 for b in BOLTS if b.status == "Parcial")
    pct = round((si + 0.5 * par) / total * 100, 1)
    print(f"Generado: {out}")
    print(f"  Hojas: 8 | Bolts: {total} | Si: {si} | No: {no} | Parcial: {par} | Avance: {pct}%")
    print(f"  Autor: {PROJECT['autor']}")


if __name__ == "__main__":
    main()
